from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from src.importer.exceptions import (
    FileMissingError,
    FileTypeNotSupportedError,
    RequiredColumnsMissingError,
)


class FileValidator:
    """Validates input files before they are processed."""

    SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
    _NORMALIZE_PATTERN = re.compile(r"[^a-z0-9]+")

    def validate(self, file_path: Path, required_columns: Iterable[str] | None = None) -> None:
        self._validate_exists(file_path)
        self._validate_extension(file_path)
        if required_columns:
            self._validate_required_columns(file_path, required_columns)

    def _validate_exists(self, file_path: Path) -> None:
        if not file_path.exists():
            raise FileMissingError(f"File not found: {file_path}")

    def _validate_extension(self, file_path: Path) -> None:
        if file_path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            raise FileTypeNotSupportedError(
                f"Unsupported file type '{file_path.suffix}' for file '{file_path}'. "
                f"Supported extensions are: {', '.join(sorted(self.SUPPORTED_EXTENSIONS))}."
            )

    def _validate_required_columns(self, file_path: Path, required_columns: Iterable[str]) -> None:
        required_columns_list = [self._normalize_column_name(column) for column in required_columns]
        actual_columns = {self._normalize_column_name(column) for column in self._read_header(file_path)}
        missing_columns = [column for column in required_columns_list if column not in actual_columns]

        if missing_columns:
            raise RequiredColumnsMissingError(list(required_columns_list), missing_columns, file_path)

    def _read_header(self, file_path: Path) -> list[str]:
        if file_path.suffix.lower() == ".csv":
            with file_path.open(newline="", encoding="utf-8") as file:
                reader = csv.reader(file)
                return next(reader, [])

        from openpyxl import load_workbook

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            header_row = next(worksheet.iter_rows(values_only=True), [])
            return [str(value) if value is not None else "" for value in header_row]
        finally:
            workbook.close()

    def _normalize_column_name(self, column_name: str) -> str:
        normalized = str(column_name or "").strip().lower()
        normalized = self._NORMALIZE_PATTERN.sub("_", normalized)
        normalized = normalized.strip("_")
        return normalized
