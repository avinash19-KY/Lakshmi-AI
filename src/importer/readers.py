from __future__ import annotations

import csv
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class Reader(ABC):
    """Reads raw records from a supported import file."""

    @abstractmethod
    def read(self, file_path: Path) -> list[dict[str, Any]]:
        raise NotImplementedError


class CsvReader(Reader):
    """Reads records from CSV files."""

    def read(self, file_path: Path) -> list[dict[str, Any]]:
        with file_path.open(newline="", encoding="utf-8") as file:
            return [dict(row) for row in csv.DictReader(file)]


class ExcelReader(Reader):
    """Reads records from Excel files."""

    def read(self, file_path: Path) -> list[dict[str, Any]]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            return []

        headers = [self._normalize_header_value(value) for value in header_row]
        records: list[dict[str, Any]] = []

        for row in rows:
            record: dict[str, Any] = {}
            for header, cell_value in zip(headers, row):
                if header:
                    record[header] = cell_value
            records.append(record)

        return records

    @staticmethod
    def _normalize_header_value(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()
